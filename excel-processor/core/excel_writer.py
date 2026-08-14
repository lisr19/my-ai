"""
Excel 输出模块
负责生成格式化的 Excel 报告，包含多个 Sheet 和样式。
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List


class ExcelWriter:
    """生成格式化的 Excel 统计报告。"""

    # 样式定义
    HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F3864")
    GROUP_FONT = Font(name="微软雅黑", size=11, bold=True)
    GROUP_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    NORMAL_FONT = Font(name="微软雅黑", size=10)
    RANK1_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # 金
    RANK2_FILL = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # 银
    RANK3_FILL = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")  # 铜
    BORDER = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center")

    def __init__(self, output_path: str):
        self.output_path = output_path
        self.wb = Workbook()

    def write_report(self, data: dict):
        """
        生成完整的统计报告 Excel。

        Args:
            data: ScoreStatistics.calculate() 返回的数据
        """
        # 删除默认 Sheet
        self.wb.remove(self.wb.active)

        # Sheet 1: 个人明细统计
        self._write_student_detail(data)

        # Sheet 2: 小组汇总统计
        self._write_group_summary(data)

        # Sheet 3: 排行榜
        self._write_ranking(data)

        # 保存
        self.wb.save(self.output_path)

    def _write_student_detail(self, data: dict):
        """Sheet 1: 个人明细统计 - 原始数据 + 个人总分 + 全班排名 + 组内排名"""
        ws = self.wb.create_sheet("个人明细统计")

        score_names = data["score_item_names"]
        students = data["students"]

        # 标题行
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6 + len(score_names))
        title_cell = ws.cell(row=1, column=1, value=data["title"])
        title_cell.font = self.TITLE_FONT
        title_cell.alignment = self.CENTER
        ws.row_dimensions[1].height = 30

        # 表头
        headers = ["组别", "姓名"] + score_names + ["个人总分", "全班排名", "组内排名"]
        header_row = 2
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.CENTER
            cell.border = self.BORDER
        ws.row_dimensions[header_row].height = 25

        # 数据行（按小组顺序，组内按排名）
        current_row = 3
        current_group = None

        # 先按小组排名排序
        sorted_students = sorted(
            students,
            key=lambda s: (s["group_rank_overall"], s["group_rank"])
        )

        for student in sorted_students:
            # 如果换组了，插入小组分隔行
            if student["group"] != current_group:
                current_group = student["group"]
                # 查找小组统计
                group_info = next(
                    g for g in data["group_summary"] if g["name"] == current_group
                )
                # 小组信息行
                ws.merge_cells(
                    start_row=current_row, start_column=1,
                    end_row=current_row, end_column=len(headers)
                )
                gcell = ws.cell(row=current_row, column=1,
                    value=f"  {current_group}  |  小组总分: {group_info['total']}  |  "
                          f"平均分: {group_info['average']}  |  小组排名: 第{group_info['rank']}名  |  "
                          f"人数: {group_info['member_count']}")
                gcell.font = self.GROUP_FONT
                gcell.fill = self.GROUP_FILL
                gcell.alignment = self.LEFT
                gcell.border = self.BORDER
                current_row += 1

            # 学生数据
            row_data = [student["group"], student["name"]] + student["scores"] + [
                student["total"],
                student["class_rank"],
                student["group_rank"]
            ]

            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.font = self.NORMAL_FONT
                cell.alignment = self.CENTER
                cell.border = self.BORDER

                # 组内排名前三名特殊颜色
            if student["group_rank"] == 1:
                ws.cell(row=current_row, column=len(headers)).fill = self.RANK1_FILL
            elif student["group_rank"] == 2:
                ws.cell(row=current_row, column=len(headers)).fill = self.RANK2_FILL
            elif student["group_rank"] == 3:
                ws.cell(row=current_row, column=len(headers)).fill = self.RANK3_FILL

            current_row += 1

        # 列宽
        ws.column_dimensions["A"].width = 12  # 组别
        ws.column_dimensions["B"].width = 12  # 姓名
        for i in range(len(score_names)):
            col_letter = get_column_letter(3 + i)
            ws.column_dimensions[col_letter].width = 10
        total_col = get_column_letter(3 + len(score_names))
        ws.column_dimensions[total_col].width = 12
        rank_col = get_column_letter(4 + len(score_names))
        ws.column_dimensions[rank_col].width = 12
        group_rank_col = get_column_letter(5 + len(score_names))
        ws.column_dimensions[group_rank_col].width = 12

        # 冻结首行
        ws.freeze_panes = "A3"

    def _write_group_summary(self, data: dict):
        """Sheet 2: 小组汇总统计"""
        ws = self.wb.create_sheet("小组汇总统计")

        # 标题
        ws.merge_cells("A1:F1")
        title_cell = ws.cell(row=1, column=1, value="小组积分汇总统计")
        title_cell.font = self.TITLE_FONT
        title_cell.alignment = self.CENTER
        ws.row_dimensions[1].height = 30

        # 表头
        headers = ["排名", "小组名称", "小组人数", "小组总分", "小组平均分", "与第一名差距"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.CENTER
            cell.border = self.BORDER
        ws.row_dimensions[2].height = 25

        # 数据
        group_summary = data["group_summary"]
        max_total = group_summary[0]["total"] if group_summary else 0

        for idx, g in enumerate(group_summary, 1):
            row = idx + 2
            gap = max_total - g["total"]

            row_data = [g["rank"], g["name"], g["member_count"], g["total"], g["average"], gap]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = self.NORMAL_FONT
                cell.alignment = self.CENTER
                cell.border = self.BORDER

            # 前三名排名着色
            if g["rank"] == 1:
                ws.cell(row=row, column=1).fill = self.RANK1_FILL
            elif g["rank"] == 2:
                ws.cell(row=row, column=1).fill = self.RANK2_FILL
            elif g["rank"] == 3:
                ws.cell(row=row, column=1).fill = self.RANK3_FILL

        # 列宽
        widths = [10, 14, 10, 12, 14, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # 冻结首行
        ws.freeze_panes = "A3"

    def _write_ranking(self, data: dict):
        """Sheet 3: 排行榜 - 全班排名 + 各组第一名"""
        ws = self.wb.create_sheet("排行榜")

        # === 全班排行榜 ===
        ws.merge_cells("A1:E1")
        title_cell = ws.cell(row=1, column=1, value="全班积分排行榜")
        title_cell.font = self.TITLE_FONT
        title_cell.alignment = self.CENTER
        ws.row_dimensions[1].height = 30

        # 表头
        headers = ["排名", "姓名", "所在小组", "个人总分", "组内排名"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.CENTER
            cell.border = self.BORDER
        ws.row_dimensions[2].height = 25

        # 全班排名数据
        sorted_students = sorted(data["students"], key=lambda x: x["total"], reverse=True)
        for idx, student in enumerate(sorted_students, 1):
            row = idx + 2
            row_data = [
                student["class_rank"],
                student["name"],
                student["group"],
                student["total"],
                f"第{student['group_rank']}名"
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = self.NORMAL_FONT
                cell.alignment = self.CENTER
                cell.border = self.BORDER

            # 前三名着色
            if student["class_rank"] == 1:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = self.RANK1_FILL
            elif student["class_rank"] == 2:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = self.RANK2_FILL
            elif student["class_rank"] == 3:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = self.RANK3_FILL

        # 列宽
        widths = [10, 12, 14, 12, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # === 各组第一名 ===
        start_row = len(sorted_students) + 5
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)
        title2 = ws.cell(row=start_row, column=1, value="各小组第一名")
        title2.font = self.TITLE_FONT
        title2.alignment = self.CENTER
        ws.row_dimensions[start_row].height = 30

        # 表头
        header_row2 = start_row + 1
        headers2 = ["小组排名", "小组名称", "第一名姓名", "个人总分", "全班排名"]
        for col, h in enumerate(headers2, 1):
            cell = ws.cell(row=header_row2, column=col, value=h)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.CENTER
            cell.border = self.BORDER
        ws.row_dimensions[header_row2].height = 25

        # 各组第一名数据
        leaders = [s for s in sorted_students if s["group_rank"] == 1]
        for idx, leader in enumerate(leaders, 1):
            row = header_row2 + idx
            row_data = [
                leader["group_rank_overall"],
                leader["group"],
                leader["name"],
                leader["total"],
                leader["class_rank"]
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = self.NORMAL_FONT
                cell.alignment = self.CENTER
                cell.border = self.BORDER

            # 小组排名第一的行着色
            if leader["group_rank_overall"] == 1:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = self.RANK1_FILL
            elif leader["group_rank_overall"] == 2:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = self.RANK2_FILL
            elif leader["group_rank_overall"] == 3:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = self.RANK3_FILL

        # 冻结首行
        ws.freeze_panes = "A3"
