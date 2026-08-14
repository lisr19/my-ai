"""
Excel 读取模块
负责读取 .xlsx 文件，解析合并单元格，提取小组和学生积分数据。
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import Optional


class ExcelReader:
    """读取 Excel 积分表，解析为结构化数据。"""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = load_workbook(file_path, data_only=True)
        self.ws = self.wb.active  # 默认使用第一个工作表

    def read_score_data(self) -> dict:
        """
        读取积分表数据。

        返回结构:
        {
            "title": "表格标题",
            "headers": ["组别", "姓名", "积分项1", ...],
            "score_item_count": 15,
            "groups": [
                {
                    "name": "第一组",
                    "members": [
                        {"name": "邵泽贤", "scores": [13, 19, 13, ...]},
                        ...
                    ]
                },
                ...
            ]
        }
        """
        # 解析标题（第1行，合并单元格）
        title = self.ws.cell(row=1, column=1).value or ""

        # 解析表头（第2行）
        # 该表中积分项列(C~Z)无表头文字，只有统计列(AA~AE)有表头
        headers = []
        header_row = 2

        for col in range(1, self.ws.max_column + 1):
            val = self.ws.cell(row=header_row, column=col).value
            if val is not None:
                headers.append((col, str(val)))

        # 找到关键列位置
        name_col = self._find_header_col(headers, "姓名")
        total_col = self._find_header_col(headers, "个总")  # 个人总分列

        # 积分项范围: 姓名列之后 ~ "个总"列之前
        score_start_col = name_col + 1
        if total_col > 0:
            score_item_end_col = total_col - 1
        else:
            # 如果没有"个总"列，使用最后一个有数据的列
            score_item_end_col = self.ws.max_column

        # 找到所有有姓名的数据行
        raw_rows = []
        for row in range(3, self.ws.max_row + 1):
            name = self.ws.cell(row=row, column=name_col).value
            if name is None or str(name).strip() == "":
                continue
            raw_rows.append((row, str(name).strip()))

        # 解析合并单元格，确定每行属于哪个小组
        group_map = self._parse_merged_groups()

        # 构建分组数据
        groups_dict = {}  # 有序字典，保持小组顺序
        group_order = []

        for row_idx, name in raw_rows:
            group_name = group_map.get(row_idx, "未分组")
            if group_name not in groups_dict:
                groups_dict[group_name] = []
                group_order.append(group_name)

            # 读取积分项
            scores = []
            for col in range(score_start_col, score_item_end_col + 1):
                val = self.ws.cell(row=row_idx, column=col).value
                if val is None:
                    scores.append(0)
                elif isinstance(val, (int, float)):
                    scores.append(val)
                else:
                    # 尝试转换为数字
                    try:
                        scores.append(float(val))
                    except (ValueError, TypeError):
                        scores.append(0)

            groups_dict[group_name].append({
                "name": name,
                "scores": scores
            })

        # 转为列表
        groups = []
        for gname in group_order:
            groups.append({
                "name": gname,
                "members": groups_dict[gname]
            })

        # 构建积分项名称列表（积分项列无表头，自动生成名称）
        score_item_names = []
        for col in range(score_start_col, score_item_end_col + 1):
            # 检查是否有自定义表头
            custom_name = None
            for h_col, h_val in headers:
                if h_col == col:
                    custom_name = h_val
                    break
            if custom_name:
                score_item_names.append(custom_name)
            else:
                score_item_names.append(f"积分项{col - score_start_col + 1}")

        return {
            "title": title,
            "score_item_names": score_item_names,
            "score_item_count": len(score_item_names),
            "groups": groups,
            "total_students": sum(len(g["members"]) for g in groups),
            "total_groups": len(groups)
        }

    def _find_header_col(self, headers: list, name: str) -> int:
        """在表头中查找指定列名对应的列号。"""
        for col, h in headers:
            if h == name:
                return col
        return -1

    def _parse_merged_groups(self) -> dict:
        """
        解析A列的合并单元格，返回 {行号: 小组名} 映射。

        合并单元格如 A3:A6 表示第3-6行都属于"第一组"。
        """
        group_map = {}

        for merge_range in self.ws.merged_cells.ranges:
            # 只处理A列的合并单元格（组别列）
            if merge_range.min_col == 1 and merge_range.max_col == 1:
                start_row = merge_range.min_row
                end_row = merge_range.max_row
                group_name = self.ws.cell(row=start_row, column=1).value
                if group_name:
                    group_name = str(group_name).strip()
                    for r in range(start_row, end_row + 1):
                        group_map[r] = group_name

        return group_map
